#!/usr/bin/env python

# Imports

from bs4 import BeautifulSoup

import requests
import re
import os
from openpyxl import Workbook, load_workbook
from datetime import date

# Program variables

temp_content_path = "./content.txt"
stock_spreadsheet_path = "./stock.xlsx"

base_url = 'https://stockx.com/nike-air-force-1-low-supreme-box-logo-white'
base_size = 9

# Request headers

headers = {
    'Host': 'stockx.com',
    'User-Agent': 'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:106.0) Gecko/20100101 Firefox/106.0',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8',
    'Referrer': 'none',
    'Accept-Language': 'en-US,en;q=0.5',
    'DNT': '1',
    'Connection': 'keep-alive',
    'Upgrade-Insecure-Requests': '1',
    'Sec-Fetch-Dest': 'document',
    'Sec-Fetch-Mode': 'navigate',
    'Sec-Fetch-Site': 'none',
    'Sec-Fetch-User': '1'
}

# Program body

def form_url():
    url = base_url + '?size=' + str(base_size)
    return url

def request_page_content(url):
    r = requests.get(url, headers=headers)
    page_content = r.text
    return page_content

def process_page_content(html):
    soup = BeautifulSoup(html, "html.parser")
    find_pid = soup.find_all("div", {"class": "chakra-container css-vp2g1e"})
    page_title = soup.find('title').text

    # Write content to file
    if not os.path.isfile(temp_content_path):
        with open(temp_content_path, "w") as f: print(find_pid, file=f)
        print("[ SYSTEM ] Creating temp file...")

    # Read content from file
    f = open(temp_content_path, "r")
    div_content = f.read()

    # Save product info
    product_brand = re.findall(r'(?<=\"Brand\",\"name\":\")(.*?)(?=\")', div_content)
    product_model = re.findall(r'(?<=\"model\":\")(.*?)(?=\")', div_content)
    product_price = re.findall(r'(?<=\"price\":)(\d+)', div_content)
    product_sku = re.findall(r'\w{6}-\w{3}', page_title)

    # Close the file
    f.close()

    # Remove the file
    if os.path.isfile(temp_content_path):
        os.remove(temp_content_path)
        print("[ SYSTEM ] Removing temp content file...")

    print(f"\nBrand: {product_brand[0]}\nModel: {product_model[0]}\nPrice: {product_price[0]}\nSKU: {product_sku[0]}\n")

    # Place product info in a list, return the list with product info
    product_info = [product_brand[0], product_model[0], product_price[0], product_sku[0]]
    return product_info

def write_spreadsheet(item_list_content):
    # Add current date to info list
    product_date = date.today().strftime("%m.%d.%y")
    item_list_content.insert(0, product_date)

    # Check if spreadsheet exists
    if not os.path.isfile(stock_spreadsheet_path):
        print("[ ALERT ]  Stock spreadsheet not found!")
        print("[ SYSTEM ] Creating stock.xlsx file...\n")

        # Define workbook
        workbook = Workbook()
        sheet = workbook.active

        # Fill sheet header
        sheet["A1"] = "Date"
        sheet["B1"] = "Brand"
        sheet["C1"] = "Model"
        sheet["D1"] = "Price"
        sheet["E1"] = "SKU"

        # Create the spreadsheet
        workbook.save(filename=stock_spreadsheet_path)
    else:
        print("[ ALERT ] Found stock.xlsx file!\n")
        print("[ SYSTEM ] Writing to file...")

        # Load workbook
        workbook = load_workbook(stock_spreadsheet_path)
        sheet = workbook.active

        # Get last row index
        maxRowSrcFile = sheet.max_row

        # Write output to last line
        iter_col = 1
        for data in item_list_content:
            sheet.cell(row=maxRowSrcFile+1, column=iter_col, value=data)
            iter_col += 1

        # Save workbook
        workbook.save(filename=stock_spreadsheet_path)
        print("[ ALERT ] Done")

def read_spreadsheet():
    # Check if spreadsheet exists
    if not os.path.isfile(stock_spreadsheet_path):
        print("[ ALERT ]  Stock spreadsheet not found!")
        print("[ SYSTEM ] Creating stock.xlsx file...\n")

        # Define workbook
        workbook = Workbook()
        sheet = workbook.active

        # Fill sheet header
        sheet["A1"] = "Date"
        sheet["B1"] = "Brand"
        sheet["C1"] = "Model"
        sheet["D1"] = "Price"
        sheet["E1"] = "SKU"

        # Write the workbook
        workbook.save(filename=stock_spreadsheet_path)
    else:
        print("[ ALERT ] Found stock.xlsx file!\n")

        # Load workbook
        workbook = load_workbook(stock_spreadsheet_path)
        sheet = workbook.active

        # Print last row & last col
        maxRowSrcFile = sheet.max_row
        maxColSrcFile = sheet.max_column
        print(f"Rows found: {maxRowSrcFile} | Cols found: {maxColSrcFile}\n")

        # Print spreadsheet content
        for value in sheet.iter_rows(values_only=True):
            print(value)

def main():
    url = form_url() # Form request URL to the product page
    html = request_page_content(url) # Send request to the URL and get page content
    info_list = process_page_content(html) # Get product info
    write_spreadsheet(info_list) # Write to spreadsheet
    read_spreadsheet() # Read spreadsheet content
    
if __name__ == '__main__':
    main()